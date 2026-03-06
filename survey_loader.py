from pathlib import Path
from typing import Dict, List
from openpyxl import load_workbook

from models import Survey, Block, Question, Option

LANGS = ("uz", "ru")


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "ha"}


def _normalize_headers(ws) -> List[str]:
    return [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]


def _row_to_dict(headers, row):
    data = {}
    for idx, header in enumerate(headers):
        data[header] = row[idx] if idx < len(row) else None
    return data


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _lang_map(data, base, fallback_key=None):
    result = {}
    for lang in LANGS:
        key = f"{base}_{lang}"
        if key in data and data[key] not in (None, ""):
            result[lang] = _clean(data[key])
    if not result and fallback_key and data.get(fallback_key) not in (None, ""):
        result["uz"] = _clean(data[fallback_key])
        result["ru"] = _clean(data[fallback_key])
    return result


def load_questionnaire(xlsx_path: Path) -> Dict[str, Survey]:
    workbook = load_workbook(xlsx_path, data_only=True)

    surveys_sheet = workbook["surveys"]
    blocks_sheet = workbook["blocks"]
    questions_sheet = workbook["questions"]
    options_sheet = workbook["options"]

    surveys: Dict[str, Survey] = {}

    headers = _normalize_headers(surveys_sheet)
    for row in surveys_sheet.iter_rows(min_row=2, values_only=True):
        data = _row_to_dict(headers, row)
        survey_code = _clean(data.get("survey_code"))
        if not survey_code:
            continue
        surveys[survey_code] = Survey(
            survey_code=survey_code,
            order_no=int(data.get("order_no") or 0),
            is_active=_to_bool(data.get("is_active")),
            titles=_lang_map(data, "title", "survey_title"),
            audiences=_lang_map(data, "audience", "audience"),
            descriptions=_lang_map(data, "description", "description"),
        )

    headers = _normalize_headers(blocks_sheet)
    for row in blocks_sheet.iter_rows(min_row=2, values_only=True):
        data = _row_to_dict(headers, row)
        survey_code = _clean(data.get("survey_code"))
        block_code = _clean(data.get("block_code"))
        if not survey_code or not block_code or survey_code not in surveys:
            continue
        surveys[survey_code].blocks[block_code] = Block(
            survey_code=survey_code,
            block_code=block_code,
            roman_no=_clean(data.get("roman_no")),
            order_no=int(data.get("order_no") or 0),
            titles=_lang_map(data, "title", "block_title"),
            intros=_lang_map(data, "description", "block_intro"),
        )

    question_index: Dict[str, Question] = {}
    headers = _normalize_headers(questions_sheet)
    for row in questions_sheet.iter_rows(min_row=2, values_only=True):
        data = _row_to_dict(headers, row)
        survey_code = _clean(data.get("survey_code"))
        question_code = _clean(data.get("question_code"))
        if not survey_code or not question_code or survey_code not in surveys:
            continue
        question = Question(
            survey_code=survey_code,
            block_code=_clean(data.get("block_code")),
            question_code=question_code,
            question_order=int(data.get("question_order") or 0),
            question_order_in_block=int(data.get("question_order_in_block") or 0),
            block_total_questions=int(data.get("block_total_questions") or 0),
            question_type=_clean(data.get("question_type")),
            required=_to_bool(data.get("required")),
            allow_skip=_to_bool(data.get("allow_skip")),
            texts=_lang_map(data, "question_text", "question_text"),
            help_texts=_lang_map(data, "help_text", "help_text"),
            placeholders=_lang_map(data, "placeholder", "placeholder"),
            validation_regex=_clean(data.get("validation_regex")),
        )
        surveys[survey_code].questions.append(question)
        question_index[question_code] = question

    headers = _normalize_headers(options_sheet)
    for row in options_sheet.iter_rows(min_row=2, values_only=True):
        data = _row_to_dict(headers, row)
        question_code = _clean(data.get("question_code"))
        if not question_code or question_code not in question_index:
            continue
        option = Option(
            order=int(data.get("option_order") or 0),
            value=_clean(data.get("option_value") or data.get("value") or data.get("option_order") or ""),
            labels=_lang_map(data, "option_text", "option_label"),
        )
        question_index[question_code].options.append(option)

    for survey in surveys.values():
        survey.questions.sort(key=lambda q: q.question_order)
        for question in survey.questions:
            question.options.sort(key=lambda o: o.order)

    return dict(sorted(surveys.items(), key=lambda item: item[1].order_no))
