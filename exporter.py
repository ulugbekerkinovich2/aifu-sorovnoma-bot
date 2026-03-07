from datetime import datetime
from pathlib import Path
from typing import Dict, List
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Survey

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _set_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def export_responses_to_excel(response_files: List[Path], surveys: Dict[str, Survey], output_path: Path) -> Path:
    responses = [json.loads(path.read_text(encoding="utf-8")) for path in response_files]

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["survey_code", "survey_title_uz", "survey_title_ru", "responses_count"])
    _set_header_style(ws)

    counts = {}
    for response in responses:
        code = response.get("survey_code", "")
        counts[code] = counts.get(code, 0) + 1

    for survey in surveys.values():
        ws.append([survey.survey_code, survey.get_title("uz"), survey.get_title("ru"), counts.get(survey.survey_code, 0)])
    ws.append([])
    ws.append(["created_at", datetime.now().isoformat(timespec="seconds")])
    _set_widths(ws, {1: 18, 2: 28, 3: 28, 4: 18})

    meta_headers = [
        "response_id", "session_id", "lang", "survey_code", "survey_title",
        "telegram_user_id", "telegram_username", "telegram_first_name", "telegram_last_name",
        "started_at", "finished_at"
    ]
    question_order = []
    question_text_map = {}
    question_type_map = {}
    for survey in surveys.values():
        for question in survey.questions:
            question_order.append(question.question_code)
            question_text_map[question.question_code] = {"uz": question.get_text("uz"), "ru": question.get_text("ru")}
            question_type_map[question.question_code] = question.question_type

    ws = wb.create_sheet("raw_wide")
    ws.append(meta_headers + question_order)
    _set_header_style(ws)
    for response in responses:
        row = [
            response.get("response_id"), response.get("session_id"), response.get("lang", "uz"),
            response.get("survey_code"), response.get("survey_title"),
            response.get("user", {}).get("telegram_user_id"), response.get("user", {}).get("username"),
            response.get("user", {}).get("first_name"), response.get("user", {}).get("last_name"),
            response.get("started_at"), response.get("finished_at"),
        ]
        answers = response.get("answers", {})
        for question_code in question_order:
            row.append(answers.get(question_code, {}).get("answer_text", ""))
        ws.append(row)
    _set_widths(ws, {1: 22, 2: 22, 3: 10, 4: 18, 5: 28, 6: 18, 7: 20, 8: 18, 9: 18, 10: 22, 11: 22})

    ws = wb.create_sheet("answers_long")
    ws.append([
        "response_id", "lang", "survey_code", "question_code",
        "question_text_uz", "question_text_ru", "question_type", "is_skipped",
        "answer_value", "answer_text", "answered_at"
    ])
    _set_header_style(ws)
    for response in responses:
        for question_code, answer in response.get("answers", {}).items():
            value = answer.get("answer_value", "")
            if isinstance(value, list):
                value = " | ".join(str(v) for v in value)
            ws.append([
                response.get("response_id"), response.get("lang", "uz"), response.get("survey_code"),
                question_code, question_text_map.get(question_code, {}).get("uz", ""),
                question_text_map.get(question_code, {}).get("ru", ""),
                question_type_map.get(question_code, ""), answer.get("skipped", False),
                value, answer.get("answer_text", ""), answer.get("answered_at", "")
            ])
    _set_widths(ws, {1: 22, 2: 10, 3: 18, 4: 18, 5: 55, 6: 55, 7: 16, 8: 12, 9: 25, 10: 45, 11: 22})

    ws = wb.create_sheet("questions_ref")
    ws.append(["survey_code", "question_code", "question_order", "question_type", "question_text_uz", "question_text_ru"])
    _set_header_style(ws)
    for survey in surveys.values():
        for question in survey.questions:
            ws.append([survey.survey_code, question.question_code, question.question_order, question.question_type, question.get_text("uz"), question.get_text("ru")])
    _set_widths(ws, {1: 18, 2: 18, 3: 14, 4: 16, 5: 55, 6: 55})

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 18:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


_RESPONSE_META_HEADERS = [
    "№", "Til", "Boshlangan vaqt", "Tugatilgan vaqt",
    "Telegram ID", "Username", "Ism-familiya",
]

_META_COL_WIDTHS = {
    "№": 5,
    "Til": 12,
    "Boshlangan vaqt": 20,
    "Tugatilgan vaqt": 20,
    "Telegram ID": 16,
    "Username": 20,
    "Ism-familiya": 26,
}

ROW_FILL_ODD = PatternFill("solid", fgColor="F2F7FF")


def _make_response_sheet(wb: Workbook, survey: Survey) -> object:
    """Create a new professional sheet for a survey with human-readable headers."""
    title = survey.get_title("uz")[:31]
    ws = wb.create_sheet(title)

    q_headers = [q.get_text("uz") or q.question_code for q in survey.questions]
    all_headers = _RESPONSE_META_HEADERS + q_headers
    ws.append(all_headers)

    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    for i, h in enumerate(_RESPONSE_META_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = _META_COL_WIDTHS.get(h, 15)
    for i in range(len(_RESPONSE_META_HEADERS) + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 35

    # Header row height for long question texts
    ws.row_dimensions[1].height = 60

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"

    return ws


def append_response_to_excel(payload: dict, surveys_dict: Dict[str, Survey], output_path: Path):
    """Append one completed response as a new row in a per-survey sheet."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    survey_code = payload.get("survey_code", "")
    survey = surveys_dict.get(survey_code)
    if not survey:
        return

    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_title = survey.get_title("uz")[:31]
    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
    else:
        ws = _make_response_sheet(wb, survey)

    row_num = ws.max_row  # header = row 1, so first data row gets №=1
    user = payload.get("user", {})
    answers = payload.get("answers", {})
    lang_display = "O'zbek" if payload.get("lang") == "uz" else "Русский"

    row = [
        row_num,
        lang_display,
        payload.get("started_at", ""),
        payload.get("finished_at", ""),
        str(user.get("telegram_user_id", "")),
        user.get("username", ""),
        user.get("full_name", ""),
    ]
    for q in survey.questions:
        ans = answers.get(q.question_code, {})
        row.append("" if ans.get("skipped") else ans.get("answer_text", ""))

    ws.append(row)

    # Alternate row shading
    data_row = ws.max_row
    if data_row % 2 == 0:
        for cell in ws[data_row]:
            cell.fill = ROW_FILL_ODD

    wb.save(output_path)
