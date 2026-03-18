from __future__ import annotations

import re
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parents[1]
UZ_DOCX = Path("/Users/m3/Desktop/so'rovnoma.docx")
RU_DOCX = Path("/Users/m3/Desktop/so_rovnoma_ru.docx")
OUTPUT_XLSX = BASE_DIR / "questionnaire" / "aifu_questionnaire_config_bilingual_full.xlsx"

SURVEY_CODES = {
    "student": {
        "uz": "TALABALAR SOROVNOMASI  (Bakalavr va Magistr)",
        "ru": "АНКЕТА ДЛЯ СТУДЕНТОВ (Бакалавр и Магистр)",
        "audience_uz": "Talabalar",
        "audience_ru": "Студенты",
    },
    "employee": {
        "uz": "XODIMLAR SOROVNOMASI",
        "ru": "АНКЕТА ДЛЯ СОТРУДНИКОВ",
        "audience_uz": "Xodimlar",
        "audience_ru": "Сотрудники",
    },
    "parent": {
        "uz": "OTA-ONALAR SOROVNOMASI",
        "ru": "АНКЕТА ДЛЯ РОДИТЕЛЕЙ",
        "audience_uz": "Ota-onalar",
        "audience_ru": "Родители",
    },
    "applicant": {
        "uz": "ABITURIYENTLAR SOROVNOMASI",
        "ru": "АНКЕТА ДЛЯ АБИТУРИЕНТОВ",
        "audience_uz": "Abituriyentlar",
        "audience_ru": "Абитуриенты",
    },
}

QUESTION_RE = re.compile(r"^(\d+)\s*/\s*(\d+)\s+(.+)$")
BLOCK_RE_UZ = re.compile(r"^([IVX]+)\s+BLOK:\s+(.+)$")
BLOCK_RE_RU = re.compile(r"^БЛОК\s+([IVX]+):\s+(.+)$")
OPTION_NUMBER_RE = re.compile(r"^\d+$")
TERMINAL_NOTE_PREFIXES = (
    "Vaqtingiz va",
    "Большое спасибо",
)


@dataclass
class ParsedOption:
    order: int
    text: str


@dataclass
class ParsedQuestion:
    order_in_block: int
    declared_total: int
    text: str
    help_text: str = ""
    options: List[ParsedOption] = field(default_factory=list)


@dataclass
class ParsedBlock:
    roman_no: str
    title: str
    intro: str
    questions: List[ParsedQuestion] = field(default_factory=list)


@dataclass
class ParsedSurvey:
    title: str
    description: str
    blocks: List[ParsedBlock] = field(default_factory=list)


def extract_lines(docx_path: Path) -> List[str]:
    output = subprocess.check_output(
        ["textutil", "-convert", "txt", "-stdout", str(docx_path)],
        text=True,
    )
    lines = []
    for raw in output.splitlines():
        line = raw.replace("\x0c", "").rstrip()
        lines.append(line)
    return lines


def is_survey_header(line: str, lang: str) -> bool:
    return any(info[lang] == line for info in SURVEY_CODES.values())


def is_block_header(line: str, lang: str) -> bool:
    pattern = BLOCK_RE_UZ if lang == "uz" else BLOCK_RE_RU
    return bool(pattern.match(line))


def is_question_header(line: str) -> bool:
    return bool(QUESTION_RE.match(line))


def next_non_empty(lines: List[str], index: int) -> int:
    while index < len(lines) and not lines[index].strip():
        index += 1
    return index


def collect_paragraph(lines: List[str], index: int, lang: str) -> tuple[str, int]:
    parts: List[str] = []
    index = next_non_empty(lines, index)
    while index < len(lines):
        line = lines[index].strip()
        if not line:
            if parts:
                break
            index += 1
            continue
        if is_survey_header(line, lang) or is_block_header(line, lang) or is_question_header(line):
            break
        parts.append(line)
        index += 1
    return " ".join(parts).strip(), index


def collect_options(body_lines: Iterable[str]) -> List[ParsedOption]:
    lines = [line.strip() for line in body_lines if line.strip()]
    options: List[ParsedOption] = []
    index = 0
    while index < len(lines):
        line = lines[index]
        if line.startswith(TERMINAL_NOTE_PREFIXES):
            break
        if not OPTION_NUMBER_RE.match(line):
            index += 1
            continue
        order = int(line)
        index += 1
        text_parts: List[str] = []
        while index < len(lines) and not OPTION_NUMBER_RE.match(lines[index]):
            if lines[index].startswith(TERMINAL_NOTE_PREFIXES):
                break
            text_parts.append(lines[index])
            index += 1
        text = " ".join(text_parts).strip()
        if text:
            options.append(ParsedOption(order=order, text=text))
    return sorted(options, key=lambda item: item.order)


def parse_doc(lines: List[str], lang: str) -> List[ParsedSurvey]:
    surveys: List[ParsedSurvey] = []
    index = 0
    while index < len(lines):
        line = lines[index].strip()
        if not is_survey_header(line, lang):
            index += 1
            continue

        survey_title = line
        survey_description, index = collect_paragraph(lines, index + 1, lang)
        survey = ParsedSurvey(title=survey_title, description=survey_description)

        while index < len(lines):
            line = lines[index].strip()
            if not line:
                index += 1
                continue
            if is_survey_header(line, lang):
                break
            if not is_block_header(line, lang):
                index += 1
                continue

            pattern = BLOCK_RE_UZ if lang == "uz" else BLOCK_RE_RU
            match = pattern.match(line)
            assert match is not None
            roman_no, block_title = match.groups()
            block_intro, index = collect_paragraph(lines, index + 1, lang)
            block = ParsedBlock(roman_no=roman_no, title=block_title, intro=block_intro)

            while index < len(lines):
                index = next_non_empty(lines, index)
                if index >= len(lines):
                    break
                current = lines[index].strip()
                if is_survey_header(current, lang) or is_block_header(current, lang):
                    break
                question_match = QUESTION_RE.match(current)
                if not question_match:
                    index += 1
                    continue

                order_in_block = int(question_match.group(1))
                declared_total = int(question_match.group(2))
                question_text = question_match.group(3).strip()
                index += 1

                body_lines: List[str] = []
                while index < len(lines):
                    current = lines[index].strip()
                    if is_survey_header(current, lang) or is_block_header(current, lang) or is_question_header(current):
                        break
                    body_lines.append(lines[index])
                    index += 1

                help_lines: List[str] = []
                remaining_lines: List[str] = []
                for body_line in body_lines:
                    stripped = body_line.strip()
                    if not stripped:
                        continue
                    if stripped.startswith("*"):
                        help_lines.append(stripped[1:].strip())
                        continue
                    if "1 dan 10 gacha" in stripped or "числа от 1 до 10" in stripped:
                        help_lines.append(stripped)
                        continue
                    remaining_lines.append(stripped)

                question = ParsedQuestion(
                    order_in_block=order_in_block,
                    declared_total=declared_total,
                    text=question_text,
                    help_text=" ".join(help_lines).strip(),
                    options=collect_options(remaining_lines),
                )
                block.questions.append(question)

            survey.blocks.append(block)

        surveys.append(survey)

    return surveys


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def fix_known_ru_text(text: str) -> str:
    replacements = {
        "Hozirgi ta'lim jarayonidan qanchalik mamnunsiz? ( 1dan 10 gacha baholang)": "Насколько вы довольны текущим учебным процессом? (оцените от 1 до 10)",
        "AIFUni do’st yoki tanishlaringizga tavsiya etasizmi?": "Порекомендуете ли вы AIFU своим друзьям или знакомым?",
        "Ish muhitidan qanchalik mamnunsiz? (10 ballik tizimda baholang)": "Насколько вы довольны рабочей атмосферой? (оцените по 10-балльной шкале)",
    }
    return replacements.get(text, text)


def infer_question_type(question_uz: ParsedQuestion, question_ru: ParsedQuestion) -> tuple[str, str]:
    help_text = f"{question_uz.help_text} {question_ru.help_text}".lower()
    text = f"{question_uz.text} {question_ru.text}".lower()
    if question_uz.options or question_ru.options:
        if "bitta variant" in help_text or "выберите один вариант" in help_text:
            return "single_choice", ""
        if "1 dan 10 gacha" in help_text or "числа от 1 до 10" in help_text:
            return "single_choice", ""
        if "bir yoki bir nechta" in help_text or "один или несколько" in help_text:
            return "multi_choice", ""
        return "single_choice", ""
    if "faqat raqam" in help_text or "только цифр" in help_text or "возраст" in text or "yosh" in text:
        return "integer", r"^\d+$"
    if "erkin yozing" in help_text or "свободный ответ" in help_text:
        return "long_text", ""
    return "text", ""


def synthesize_scale_options(question_uz: ParsedQuestion, question_ru: ParsedQuestion) -> None:
    combined = f"{question_uz.help_text} {question_ru.help_text}".lower()
    if question_uz.options or question_ru.options:
        return
    if "1 dan 10 gacha" not in combined and "числа от 1 до 10" not in combined:
        return
    scale = [ParsedOption(order=value, text=str(value)) for value in range(1, 11)]
    question_uz.options = list(scale)
    question_ru.options = list(scale)


def validate_parallel_structure(uz_surveys: List[ParsedSurvey], ru_surveys: List[ParsedSurvey]) -> None:
    if len(uz_surveys) != len(ru_surveys):
        raise ValueError("UZ va RU survey soni mos emas.")
    for uz_survey, ru_survey in zip(uz_surveys, ru_surveys):
        if len(uz_survey.blocks) != len(ru_survey.blocks):
            raise ValueError(f"Block soni mos emas: {uz_survey.title}")
        for uz_block, ru_block in zip(uz_survey.blocks, ru_survey.blocks):
            if len(uz_block.questions) != len(ru_block.questions):
                raise ValueError(f"Savol soni mos emas: {uz_survey.title} / {uz_block.title}")


def build_workbook(uz_surveys: List[ParsedSurvey], ru_surveys: List[ParsedSurvey]) -> Workbook:
    validate_parallel_structure(uz_surveys, ru_surveys)
    wb = Workbook()
    wb.remove(wb.active)

    surveys_ws = wb.create_sheet("surveys")
    blocks_ws = wb.create_sheet("blocks")
    questions_ws = wb.create_sheet("questions")
    options_ws = wb.create_sheet("options")

    surveys_ws.append(
        [
            "survey_code",
            "title_uz",
            "title_ru",
            "audience_uz",
            "audience_ru",
            "order_no",
            "description_uz",
            "description_ru",
            "is_active",
        ]
    )
    blocks_ws.append(
        [
            "survey_code",
            "block_code",
            "roman_no",
            "order_no",
            "title_uz",
            "title_ru",
            "description_uz",
            "description_ru",
        ]
    )
    questions_ws.append(
        [
            "survey_code",
            "block_code",
            "question_code",
            "question_order",
            "question_order_in_block",
            "block_total_questions",
            "question_text_uz",
            "question_text_ru",
            "help_text_uz",
            "help_text_ru",
            "question_type",
            "required",
            "allow_skip",
            "validation_regex",
            "placeholder_uz",
            "placeholder_ru",
        ]
    )
    options_ws.append(
        [
            "survey_code",
            "question_code",
            "option_order",
            "option_value",
            "option_text_uz",
            "option_text_ru",
        ]
    )

    survey_order = list(SURVEY_CODES.keys())

    for order_no, (survey_code, uz_survey, ru_survey) in enumerate(
        zip(survey_order, uz_surveys, ru_surveys), start=1
    ):
        survey_info = SURVEY_CODES[survey_code]
        surveys_ws.append(
            [
                survey_code,
                normalize_text(uz_survey.title),
                normalize_text(ru_survey.title),
                survey_info["audience_uz"],
                survey_info["audience_ru"],
                order_no,
                normalize_text(uz_survey.description),
                normalize_text(ru_survey.description),
                True,
            ]
        )

        question_order = 1
        for block_index, (uz_block, ru_block) in enumerate(zip(uz_survey.blocks, ru_survey.blocks), start=1):
            block_code = f"{survey_code}_b{block_index:02d}"
            blocks_ws.append(
                [
                    survey_code,
                    block_code,
                    uz_block.roman_no,
                    block_index,
                    normalize_text(uz_block.title),
                    normalize_text(ru_block.title),
                    normalize_text(uz_block.intro),
                    normalize_text(ru_block.intro),
                ]
            )

            block_total = len(uz_block.questions)
            for question_index, (uz_question, ru_question) in enumerate(
                zip(uz_block.questions, ru_block.questions), start=1
            ):
                synthesize_scale_options(uz_question, ru_question)
                if len(uz_question.options) != len(ru_question.options):
                    raise ValueError(
                        f"Option soni mos emas: {survey_code} / {block_code} / {question_index}"
                    )

                question_code = f"{survey_code}_q{question_order:03d}"
                question_type, validation_regex = infer_question_type(uz_question, ru_question)

                questions_ws.append(
                    [
                        survey_code,
                        block_code,
                        question_code,
                        question_order,
                        question_index,
                        block_total,
                        normalize_text(uz_question.text),
                        normalize_text(fix_known_ru_text(ru_question.text)),
                        normalize_text(uz_question.help_text),
                        normalize_text(ru_question.help_text),
                        question_type,
                        False,
                        True,
                        validation_regex,
                        "",
                        "",
                    ]
                )

                for uz_option, ru_option in zip(uz_question.options, ru_question.options):
                    options_ws.append(
                        [
                            survey_code,
                            question_code,
                            uz_option.order,
                            str(uz_option.order),
                            normalize_text(uz_option.text),
                            normalize_text(ru_option.text),
                        ]
                    )

                question_order += 1

    return wb


def main() -> None:
    uz_surveys = parse_doc(extract_lines(UZ_DOCX), "uz")
    ru_surveys = parse_doc(extract_lines(RU_DOCX), "ru")
    workbook = build_workbook(uz_surveys, ru_surveys)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    total_questions = workbook["questions"].max_row - 1
    total_options = workbook["options"].max_row - 1
    print(f"Saved {OUTPUT_XLSX}")
    print(f"Questions: {total_questions}")
    print(f"Options: {total_options}")


if __name__ == "__main__":
    main()
